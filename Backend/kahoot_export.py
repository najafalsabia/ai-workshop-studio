"""
Fills Kahoot's OWN official spreadsheet template with a generated quiz's
questions, ready to import at create.kahoot.it (Create -> Add question ->
Import spreadsheet). This does NOT build a spreadsheet from scratch — it
opens Kahoot's real downloaded template and writes into it, because Kahoot
requires their own template's internal structure to accept an import.

No LLM calls here — pure post-processing, same role as notebook_builder.py
(labs) and quiz_doc_builder.py (Word export).

Kahoot's template structure (confirmed from the real downloaded file):
  - Row 8: headers (Question, Answer 1-4, Time limit, Correct answer(s))
  - Rows 9-108: 100 pre-numbered question slots
  - Column A: question number (already filled in by Kahoot, left alone)
  - Column B: question text — max 120 characters
  - Columns C-F: Answer 1-4 — max 75 characters each
  - Column G: time limit in seconds — must be EXACTLY one of
    5, 10, 20, 30, 60, 90, 120, 240 (Kahoot rejects any other value)
  - Column H: correct answer(s) — the NUMBER(S) of the correct option(s)
    (1-4), comma-separated if more than one — NOT the answer text itself
"""

import openpyxl

from config import (
    KAHOOT_TEMPLATE_PATH,
    KAHOOT_MAX_QUESTIONS as MAX_QUESTIONS,
    KAHOOT_TIME_LIMIT_BY_DIFFICULTY as TIME_LIMIT_BY_DIFFICULTY,
    KAHOOT_DEFAULT_TIME_LIMIT as DEFAULT_TIME_LIMIT,
)

HEADER_ROW = 8
FIRST_DATA_ROW = 9

QUESTION_MAX_LEN = 120
ANSWER_MAX_LEN = 75
VALID_TIME_LIMITS = (5, 10, 20, 30, 60, 90, 120, 240)


def export_quiz_to_kahoot_xlsx(
    quiz_result: dict,
    template_path: str = str(KAHOOT_TEMPLATE_PATH),
    output_path: str = "kahoot_import.xlsx",
) -> str:
    """
    Takes generate_quiz's output ({"quiz": {"title": ..., "questions": [...]}})
    and writes it into a COPY of Kahoot's own template, ready to upload
    at create.kahoot.it. Returns the output file path.

    template_path must point to Kahoot's actual downloaded template file
    (the one with "Quiz template" in cell B2) — this function does not
    (and cannot reliably) build a compatible file from scratch.

    Raises clearly, before writing anything unrecoverable, if:
      - there are more than 100 questions (the template's limit)
      - any question doesn't have exactly 4 options
      - a question's correct_answer doesn't match any of its 4 options
        (Kahoot would have no way to know which one is correct)

    Truncates (and warns about) any question/answer text over Kahoot's
    character limits, rather than silently letting Kahoot's own importer
    reject the whole row later.
    """
    quiz = quiz_result["quiz"]
    questions = quiz.get("questions", [])

    if len(questions) > MAX_QUESTIONS:
        raise ValueError(
            f"Kahoot's template only has {MAX_QUESTIONS} question rows, but this "
            f"quiz has {len(questions)} questions. Trim the quiz before exporting."
        )

    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["Sheet1"]

    warnings = []

    for i, q in enumerate(questions):
        row = FIRST_DATA_ROW + i
        q_num = i + 1

        question_text = q.get("question", "")
        if len(question_text) > QUESTION_MAX_LEN:
            warnings.append(
                f"Q{q_num}: question was {len(question_text)} chars, truncated to {QUESTION_MAX_LEN}."
            )
            question_text = question_text[:QUESTION_MAX_LEN]
        sheet.cell(row=row, column=2, value=question_text)  # column B

        options = q.get("options", [])
        if len(options) != 4:
            raise ValueError(
                f"Q{q_num} has {len(options)} options, Kahoot's template needs exactly 4. "
                f"Options were: {options!r}"
            )

        correct_answer = q.get("correct_answer")
        correct_option_numbers = []
        for option_index, option_text in enumerate(options):
            column = 3 + option_index  # C, D, E, F for options 1-4
            text = option_text
            if len(text) > ANSWER_MAX_LEN:
                warnings.append(
                    f"Q{q_num} option {option_index + 1}: was {len(text)} chars, truncated to {ANSWER_MAX_LEN}."
                )
                text = text[:ANSWER_MAX_LEN]
            sheet.cell(row=row, column=column, value=text)
            if option_text == correct_answer:
                correct_option_numbers.append(str(option_index + 1))

        if not correct_option_numbers:
            raise ValueError(
                f"Q{q_num}: correct_answer {correct_answer!r} doesn't exactly match any of "
                f"the 4 options {options!r} — Kahoot needs to know which option number is correct."
            )

        time_limit = TIME_LIMIT_BY_DIFFICULTY.get(q.get("difficulty"), DEFAULT_TIME_LIMIT)
        sheet.cell(row=row, column=7, value=time_limit)  # column G
        sheet.cell(row=row, column=8, value=",".join(correct_option_numbers))  # column H

    workbook.save(output_path)

    if warnings:
        print(f"⚠️  {len(warnings)} warning(s) while exporting to Kahoot format:")
        for w in warnings:
            print("  -", w)

    return output_path


if __name__ == "__main__":
    # Self-test — fills a real copy of the actual Kahoot template (no API
    # keys/network needed) and re-reads it with openpyxl to confirm every
    # cell landed exactly where Kahoot expects it.

    import os

    mock_quiz_result = {
        "quiz": {
            "title": "Beyond Syntax — Final Quiz",
            "questions": [
                {
                    "question": "What is the primary risk of ambiguous prompts?",
                    "options": ["Slower response time", "Unreliable output", "Higher cost", "Shorter answers"],
                    "correct_answer": "Unreliable output",
                    "difficulty": "easy",
                },
                {
                    "question": "Which step comes first in the Trust-But-Verify pipeline?",
                    "options": ["Manual audit", "Syntax linting", "Deployment", "Documentation"],
                    "correct_answer": "Syntax linting",
                    "difficulty": "hard",
                },
            ],
        }
    }

    template = "KahootQuizTemplate.xlsx"
    output = "/tmp/test_kahoot_import.xlsx"

    print("=== Test 1: export runs and writes a file ===")
    path = export_quiz_to_kahoot_xlsx(mock_quiz_result, template_path=template, output_path=output)
    assert os.path.isfile(path)
    print(f"saved to: {path}")

    print("\n=== Test 2: re-read the file — every cell lands where Kahoot expects ===")
    wb = openpyxl.load_workbook(output)
    ws = wb["Sheet1"]

    row9 = [ws.cell(row=9, column=c).value for c in range(1, 9)]
    print("Row 9 (Q1):", row9)
    assert row9[1] == "What is the primary risk of ambiguous prompts?"  # B
    assert row9[2] == "Slower response time"  # C
    assert row9[3] == "Unreliable output"  # D
    assert row9[6] == 20  # G — easy = 20s
    assert row9[7] == "2"  # H — "Unreliable output" is option 2

    row10 = [ws.cell(row=10, column=c).value for c in range(1, 9)]
    print("Row 10 (Q2):", row10)
    assert row10[6] == 60  # G — hard = 60s
    assert row10[7] == "2"  # H — "Syntax linting" is option 2

    print("\n=== Test 3: correct_answer that doesn't match any option raises clearly ===")
    bad_quiz = {
        "quiz": {
            "questions": [
                {"question": "Q", "options": ["a", "b", "c", "d"], "correct_answer": "z", "difficulty": "easy"}
            ]
        }
    }
    try:
        export_quiz_to_kahoot_xlsx(bad_quiz, template_path=template, output_path="/tmp/test_bad.xlsx")
        raise AssertionError("Should have raised ValueError")
    except ValueError as e:
        print("Correctly rejected:", e)

    print("\n=== Test 4: more than 100 questions raises clearly ===")
    huge_quiz = {
        "quiz": {
            "questions": [
                {"question": f"Q{i}", "options": ["a", "b", "c", "d"], "correct_answer": "a", "difficulty": "easy"}
                for i in range(101)
            ]
        }
    }
    try:
        export_quiz_to_kahoot_xlsx(huge_quiz, template_path=template, output_path="/tmp/test_huge.xlsx")
        raise AssertionError("Should have raised ValueError")
    except ValueError as e:
        print("Correctly rejected:", e)

    print("\n=== Test 5: a too-long question gets truncated with a warning, not silently mangled ===")
    long_quiz = {
        "quiz": {
            "questions": [
                {
                    "question": "X" * 150,
                    "options": ["a" * 80, "b", "c", "d"],
                    "correct_answer": "a" * 80,
                    "difficulty": "medium",
                }
            ]
        }
    }
    export_quiz_to_kahoot_xlsx(long_quiz, template_path=template, output_path="/tmp/test_long.xlsx")
    wb2 = openpyxl.load_workbook("/tmp/test_long.xlsx")
    ws2 = wb2["Sheet1"]
    q_len = len(ws2.cell(row=9, column=2).value)
    a_len = len(ws2.cell(row=9, column=3).value)
    print(f"question length after truncation: {q_len} (expect <= {QUESTION_MAX_LEN})")
    print(f"answer length after truncation: {a_len} (expect <= {ANSWER_MAX_LEN})")
    assert q_len <= QUESTION_MAX_LEN
    assert a_len <= ANSWER_MAX_LEN

    print("\nAll self-tests passed.")
